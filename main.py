from time import sleep
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager


def configurar_driver():
    options = webdriver.ChromeOptions()
    options.add_argument('--no-sandbox')
    options.add_argument('--disable-dev-shm-usage')
    options.add_argument('--log-level=3')
    service = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=service, options=options)
    return driver


def formatar_preco(preco_str):
    try:
        preco_str = preco_str.replace("R$", "").replace(".", "").replace(",", ".").strip()
        return float(preco_str)
    except:
        return float('inf')


def rolar_ate_o_fim(driver, pausas=3, intervalo=1.5):
    from selenium.webdriver.common.action_chains import ActionChains
    from selenium.webdriver.common.keys import Keys
    for _ in range(pausas):
        ActionChains(driver).send_keys(Keys.END).perform()
        sleep(intervalo)


def raspar_dados_google_shopping(query):
    driver = configurar_driver()
    products_list = []

    try:
        result = query.replace(' ', '+')
        search_link = f'https://www.google.com/search?sca_esv=1ed5a24754748f85&sxsrf=AHTn8zo8SpuhsRD8V4mdHbpXO_WGW68nYQ:1744220546116&q={result}&tbm=shop&source=lnms&fbs=ABzOT_BYhiZpMrUAF0c9tORwPGlsASvANxUN_4u1oltdAlXXukJgrc8Sd9VQnu1m4CeFWCV1NFbj-Y0EivjyBcIM3oBQwXKIHetBIfJSp9D7p-3kxBrpagoTQLztGa6Nv5ZlDjulzpvQmDkY4VJEKHUsGoTOk5ZFfdzIyx0T9bFBg5uzqhjsmuJBmIIgKujST8BzF4iQUzzkQIdPbEd1BPOIHikxVCDFdA&ved=1t:200715&ictx=111&biw=1024&bih=687&dpr=1'
        driver.get(search_link)
        sleep(5)

        for pagina in range(2):
            print(f"\n📄 Página {pagina + 1}")
            rolar_ate_o_fim(driver, pausas=5, intervalo=1.5)
            sleep(3)

            products = driver.find_elements(By.CSS_SELECTOR, "div.sh-dgr__grid-result")
            if not products:
                print("⚠️ Nenhum produto encontrado nesta página.")

            for produto in products:
                try:
                    titulo = produto.find_element(By.CSS_SELECTOR, "div.EI11Pd").text
                except:
                    titulo = "Não encontrado"

                try:
                    preco = produto.find_element(By.CSS_SELECTOR, "span.a8Pemb").text
                except:
                    preco = "Não encontrado"

                try:
                    loja = produto.find_element(By.CSS_SELECTOR, "div.aULzUe").text
                except:
                    loja = "Não encontrado"

                try:
                    link = produto.find_element(By.CSS_SELECTOR, "a.shntl").get_attribute("href")
                except:
                    link = "Não encontrado"

                try:
                    reviews = produto.find_element(By.CSS_SELECTOR, "span.QIrs8").text
                except:
                    reviews = "-"

                _product = {
                    "titulo": titulo,
                    "preco": preco,
                    "loja": loja,
                    "link": link,
                    "reviews": reviews,
                }
                products_list.append(_product)

            try:
                next_btn = driver.find_element(By.CSS_SELECTOR, f"a.fl[aria-label='Page {pagina + 2}']")
                next_btn.click()
                sleep(3)
            except:
                print("⚠️ Botão 'Próxima' não encontrado ou indisponível. Parando aqui.")


    finally:
        print(f'Total de produtos raspados: {len(products_list)}')
        driver.quit()
        return products_list


def exibir_produtos_ordenados(produtos):
    for i, product in enumerate(produtos, 1):
        print(f"{i} - Título: {product.get('titulo')}")
        print(f"Preço: {product.get('preco')}")
        print(f"Reviews: {product.get('reviews')}")
        print(f"Loja:: {product.get('loja')}")
        print(f"URL: {product.get('link')}\n")


def salvar_em_excel(produtos, nome_arquivo):
    arquivo_excel = f"{nome_arquivo}.xlsx"
    df = pd.DataFrame(produtos)
    df.to_excel(arquivo_excel, index=False)

    wb = load_workbook(arquivo_excel)
    ws = wb.active

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    coluna_link_idx = None
    for i, col in enumerate(df.columns):
        if col.lower() == 'link':
            coluna_link_idx = i + 1
            break

    if coluna_link_idx:
        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=coluna_link_idx)
            url = cell.value
            if url:
                cell.value = f'=HYPERLINK("{url}", "Visitar site")'
                cell.font = Font(color="0000FF", underline="single")

    wb.save(arquivo_excel)
    print(f"✅ Planilha criada com sucesso: {arquivo_excel}")


def _search_scrapper():
    query = input("Digite o que deseja pesquisar no Google Shopping: ")
    produtos = raspar_dados_google_shopping(query)

    if not produtos:
        print("Nenhum produto encontrado, operação cancelada.")
        return

    for product in produtos:
        product['preco_num'] = formatar_preco(product.get('preco', ''))

    produtos.sort(key=lambda x: x['preco_num'])

    exibir_produtos_ordenados(produtos)
    plot_lojas_com_mais_resultados(produtos)
    nome_arquivo = input("Digite o nome do arquivo para salvar: ")
    if nome_arquivo.strip():
        salvar_em_excel(produtos, nome_arquivo)
    else:
        print("Nome de arquivo não fornecido. Salvamento cancelado.")


def plot_lojas_com_mais_resultados(produtos):
    lojas = {}
    for produto in produtos:
        loja = produto['loja']
        if loja in lojas:
            lojas[loja] += 1
        else:
            lojas[loja] = 1

    lojas = dict(sorted(lojas.items(), key=lambda item: item[1], reverse=True)[:10])

    plt.figure(figsize=(12, 8))
    plt.barh(list(lojas.keys())[::-1], list(lojas.values())[::-1])
    plt.xlabel('Quantidade de Resultados')
    plt.title('Lojas com Mais Resultados no Google Shopping')
    plt.tight_layout()
    plt.savefig('output.png')


def main():
    while True:
        print(''' 
         _______
        |.-----.|
        ||x . x||
        ||_.-._||
        `--)-(--`
       __[=== o]___
      |:::::::::::|
         ''')
        print("1 - Realizar Busca")
        print("2 - Sair")
        _response = input("Escolha uma opção:")
        if _response == "1":
            _search_scrapper()
        elif _response == "2":
            break


if __name__ == '__main__':
    main()
